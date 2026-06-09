"""
Лицензия PDF -> координаты ГК ГСК-2011 -> Excel
Использование: python3 license_to_gk.py <путь_к_pdf> [номер_зоны]
"""

import sys
import re
import pdfplumber
import pandas as pd
from pyproj import Transformer
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter


# ─── Парсинг PDF ──────────────────────────────────────────────────────────────

def parse_license_pdf(pdf_path: str) -> list[tuple]:
    """
    Извлекает таблицу угловых точек из PDF лицензии Роснедра.
    Возвращает список кортежей: (№, lat_d, lat_m, lat_s, lon_d, lon_m, lon_s)
    """
    rows = []

    with pdfplumber.open(pdf_path) as pdf:
        for page in pdf.pages:
            for table in page.extract_tables():
                for row in table:
                    if not row or not row[0]:
                        continue
                    if not re.match(r'^\d+$', str(row[0]).strip()):
                        continue
                    if len(row) < 7 or any(r is None for r in row[:7]):
                        continue
                    try:
                        pt    = int(row[0])
                        lat_d = int(row[1])
                        lat_m = int(row[2])
                        lat_s = float(str(row[3]).replace(',', '.'))
                        lon_d = int(row[4])
                        lon_m = int(row[5])
                        lon_s = float(str(row[6]).replace(',', '.'))
                        rows.append((pt, lat_d, lat_m, lat_s, lon_d, lon_m, lon_s))
                    except (ValueError, TypeError):
                        pass

    rows.sort(key=lambda x: x[0])
    return rows


# ─── Геодезия ─────────────────────────────────────────────────────────────────

def dms_to_dd(d: int, m: int, s: float) -> float:
    return d + m / 60 + s / 3600


def detect_zone(points: list[tuple]) -> int:
    avg_lon = sum(dms_to_dd(r[4], r[5], r[6]) for r in points) / len(points)
    return int(avg_lon / 6) + 1


def convert_to_gk(points: list[tuple], zone: int | None = None) -> pd.DataFrame:
    """
    Конвертирует точки ГСК-2011 (ГМС) -> ГК ГСК-2011 (метры).
    zone: номер зоны (определяется автоматически если None)
    """
    if zone is None:
        zone = detect_zone(points)

    epsg = 20900 + zone  # EPSG:20910 = GSK-2011 / GK zone 10
    transformer = Transformer.from_crs("EPSG:7683", f"EPSG:{epsg}", always_xy=True)

    records = []
    for pt, ld, lm, ls, od, om, os_ in points:
        lat = dms_to_dd(ld, lm, ls)
        lon = dms_to_dd(od, om, os_)
        easting, northing = transformer.transform(lon, lat)
        records.append({
            "№ точки":         pt,
            "Широта (°'\")" :  f"{ld}°{lm}'{ls}\"",
            "Долгота (°'\")":  f"{od}°{om}'{os_}\"",
            "Широта (DD)":     round(lat, 8),
            "Долгота (DD)":    round(lon, 8),
            "X, север (м)":    round(northing, 3),
            "Y, восток (м)":   round(easting, 3),
        })

    df = pd.DataFrame(records)
    df.attrs["zone"] = zone
    df.attrs["epsg"] = epsg
    return df


# ─── Excel ────────────────────────────────────────────────────────────────────

def save_excel(df: pd.DataFrame, out_path: str) -> None:
    zone = df.attrs.get("zone", "?")
    epsg = df.attrs.get("epsg", "?")

    with pd.ExcelWriter(out_path, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="Координаты ГК", startrow=3)
        ws = writer.sheets["Координаты ГК"]

        # Заголовок
        ws["A1"] = "Координаты угловых точек участка недр"
        ws["A1"].font = Font(bold=True, size=13)
        ws["A2"] = f"Система координат: ГСК-2011, проекция Гаусса-Крюгера, зона {zone} (EPSG:{epsg})"
        ws["A2"].font = Font(italic=True, size=10)

        # Стиль шапки таблицы (строка 4)
        header_fill = PatternFill("solid", fgColor="1F4E79")
        header_font = Font(bold=True, color="FFFFFF", size=10)
        thin = Side(style="thin")
        border = Border(left=thin, right=thin, top=thin, bottom=thin)

        for cell in ws[4]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", wrap_text=True)
            cell.border = border

        # Стиль данных
        for row in ws.iter_rows(min_row=5, max_row=4 + len(df)):
            for i, cell in enumerate(row):
                cell.border = border
                cell.alignment = Alignment(horizontal="center")
                # Чередование строк
                if (cell.row - 5) % 2 == 0:
                    cell.fill = PatternFill("solid", fgColor="EBF3FB")

        # Ширина колонок
        col_widths = [10, 18, 18, 14, 14, 16, 16]
        for i, w in enumerate(col_widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = w

        ws.row_dimensions[4].height = 30


# ─── Main ─────────────────────────────────────────────────────────────────────

def main():
    pdf_path = sys.argv[1] if len(sys.argv) > 1 else None
    zone_arg = int(sys.argv[2]) if len(sys.argv) > 2 else None

    if not pdf_path:
        print("Использование: python3 license_to_gk.py <путь_к_pdf> [зона]")
        sys.exit(1)

    print(f"📄 Читаем: {pdf_path}")
    points = parse_license_pdf(pdf_path)
    print(f"✅ Найдено точек: {len(points)}")

    df = convert_to_gk(points, zone=zone_arg)
    zone = df.attrs["zone"]
    epsg = df.attrs["epsg"]
    print(f"📐 Зона ГК: {zone} (EPSG:{epsg})")
    print(df[["№ точки", "X, север (м)", "Y, восток (м)"]].to_string(index=False))

    out = pdf_path.replace(".pdf", "_GK.xlsx")
    save_excel(df, out)
    print(f"\n💾 Сохранено: {out}")
    return df


if __name__ == "__main__":
    main()
